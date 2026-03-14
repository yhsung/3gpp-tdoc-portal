"""Real 3GPP TDoc fetching, local artifact scanning, and download integration."""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path
from typing import TYPE_CHECKING
from urllib.parse import urljoin

import openpyxl
import requests
import streamlit as st
from bs4 import BeautifulSoup

if TYPE_CHECKING:
    from portal.mock_data import TDoc

PROJECT_ROOT = Path(__file__).parent.parent
ARTIFACTS_DIR = PROJECT_ROOT / "artifacts"
DOWNLOAD_DIR = ARTIFACTS_DIR / "tdocs"
EXTRACT_DIR = ARTIFACTS_DIR / "extracted"
MD_DIR = ARTIFACTS_DIR / "output" / "markdown"
HTML_DIR = ARTIFACTS_DIR / "output" / "html"

TDOC_PATTERN = re.compile(r"(R\d-\d{7})\.zip", re.IGNORECASE)
MEETING_PATTERN = re.compile(r"TSGR1_(\d+[a-z]*)", re.IGNORECASE)


def infer_meeting_from_url(url: str) -> str:
    """Parse meeting name from a 3GPP FTP URL.

    Examples:
        .../TSGR1_120/Docs/ → "RAN1 #120"
        .../TSGR1_119bis/Docs/ → "RAN1 #119bis"
    """
    match = MEETING_PATTERN.search(url)
    if match:
        return f"RAN1 #{match.group(1)}"
    return "RAN1"


def fetch_tdoc_ids(url: str, timeout: int = 30) -> list[str]:
    """Scrape a 3GPP FTP directory listing and return TDoc IDs (e.g. 'R1-2501001')."""
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    ids: list[str] = []
    for link in soup.find_all("a", href=True):
        match = TDOC_PATTERN.search(link["href"])
        if match:
            ids.append(match.group(1))
    return ids


def _extract_title_from_md(md_path: Path) -> str:
    """Return the first H1 heading found in a markdown file, or empty string."""
    try:
        for line in md_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            stripped = line.strip()
            if stripped.startswith("# "):
                return stripped[2:].strip()
    except OSError:
        pass
    return ""


def scan_local_artifacts() -> dict[str, dict]:
    """Return a mapping of TDoc ID → artifact info for locally converted documents.

    Scans artifacts/output/markdown/ for files named ``R?-XXXXXXX_*.md``.
    """
    result: dict[str, dict] = {}
    if not MD_DIR.exists():
        return result
    for md_file in MD_DIR.glob("*.md"):
        # Expected pattern: R1-XXXXXXX_sourcefile.md
        stem = md_file.stem
        sep = stem.find("_")
        if sep < 0:
            continue
        tdoc_id = stem[:sep]
        if not re.match(r"R\d-\d{7}$", tdoc_id):
            continue
        source_file = stem[sep + 1:]
        html_file = HTML_DIR / f"{stem}.html"
        # A TDoc may have multiple output files; keep the first one found.
        if tdoc_id not in result:
            result[tdoc_id] = {
                "md_path": str(md_file),
                "html_path": str(html_file),
                "source_file": source_file,
                "title": _extract_title_from_md(md_file),
            }
    return result


@st.cache_data(ttl=3600, show_spinner=False)
def _cached_fetch_tdoc_ids(url: str) -> list[str]:
    return fetch_tdoc_ids(url)


def fetch_tdoc_metadata(docs_url: str, timeout: int = 30) -> dict[str, dict]:
    """Find and parse a TDoc list Excel from the meeting root directory.

    Walks up from the Docs/ URL to the meeting root, looks for an ``*.xlsx``
    file whose name contains "tdoc", downloads it, and extracts per-TDoc
    metadata (title, agenda_item, source).

    Returns an empty dict if no suitable file is found or parsing fails.
    """
    # Derive meeting root URL from the Docs/ URL
    root_url = docs_url.rstrip("/")
    if root_url.lower().endswith("/docs"):
        root_url = root_url[:-5]
    root_url = root_url.rstrip("/") + "/"

    try:
        resp = requests.get(root_url, timeout=timeout)
        resp.raise_for_status()
    except Exception:
        return {}

    soup = BeautifulSoup(resp.text, "html.parser")
    list_url = None
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if href.lower().endswith(".xlsx") and "tdoc" in href.lower():
            list_url = urljoin(root_url, href)
            break

    if not list_url:
        return {}

    try:
        resp = requests.get(list_url, timeout=timeout)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception:
        return {}

    ws = wb.active
    col_id = col_title = col_agenda = col_source = None
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        lowered = [str(c).lower().strip() if c is not None else "" for c in row]
        if any("tdoc" in v or "number" in v for v in lowered):
            header_row_idx = row_idx
            for j, v in enumerate(lowered):
                if ("tdoc" in v or "number" in v) and col_id is None:
                    col_id = j
                elif "title" in v and col_title is None:
                    col_title = j
                elif "agenda" in v and col_agenda is None:
                    col_agenda = j
                elif "source" in v and col_source is None:
                    col_source = j
            break

    if col_id is None or col_title is None or header_row_idx is None:
        return {}

    metadata: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if not row or not row[col_id]:
            continue
        tdoc_id = str(row[col_id]).strip()
        if not re.match(r"R\d-\d{6,7}$", tdoc_id):
            continue
        metadata[tdoc_id] = {
            "title": str(row[col_title]).strip() if row[col_title] else tdoc_id,
            "agenda_item": str(row[col_agenda]).strip() if col_agenda is not None and row[col_agenda] else "—",
            "source": str(row[col_source]).strip() if col_source is not None and row[col_source] else "",
        }

    return metadata


@st.cache_data(ttl=3600, show_spinner=False)
def _cached_fetch_tdoc_metadata(url: str) -> dict[str, dict]:
    return fetch_tdoc_metadata(url)


def get_real_documents(url: str) -> list["TDoc"]:
    """Fetch real TDoc list from a 3GPP FTP URL and return TDoc objects.

    Documents that have been locally downloaded and converted are marked
    ``available=True`` with correct ``md_path``/``html_path``/``source_file``.
    Others are stub entries with just the ID and placeholder metadata.
    """
    from portal.mock_data import TDoc  # local import to avoid circular dependency

    tdoc_ids = _cached_fetch_tdoc_ids(url)
    meeting = infer_meeting_from_url(url)
    local = scan_local_artifacts()
    metadata = _cached_fetch_tdoc_metadata(url)

    docs: list[TDoc] = []
    for tdoc_id in tdoc_ids:
        meta = metadata.get(tdoc_id, {})
        artifact = local.get(tdoc_id)
        title = meta.get("title") or (artifact or {}).get("title") or tdoc_id
        agenda_item = meta.get("agenda_item", "—")
        if artifact:
            doc = TDoc(
                id=tdoc_id,
                title=title,
                source_file=artifact["source_file"],
                file_type="docx",
                meeting=meeting,
                agenda_item=agenda_item,
                available=True,
                html_path=artifact["html_path"],
                md_path=artifact["md_path"],
            )
        else:
            doc = TDoc(
                id=tdoc_id,
                title=title,
                source_file=tdoc_id,
                file_type="zip",
                meeting=meeting,
                agenda_item=agenda_item,
            )
        docs.append(doc)
    return docs


def refresh_local_availability(docs: list) -> None:
    """Update availability, paths, and title in-place for locally converted docs.

    Called on every Streamlit rerun so the left column always reflects the
    current state of the artifacts directory without a full network re-fetch.
    """
    local = scan_local_artifacts()
    for doc in docs:
        artifact = local.get(doc.id)
        if artifact:
            doc.available = True
            doc.md_path = artifact["md_path"]
            doc.html_path = artifact["html_path"]
            doc.source_file = artifact["source_file"]
            if artifact.get("title") and doc.title == doc.id:
                doc.title = artifact["title"]


def _ensure_artifact_dirs() -> None:
    for d in (DOWNLOAD_DIR, EXTRACT_DIR, MD_DIR, HTML_DIR):
        d.mkdir(parents=True, exist_ok=True)


def download_and_convert_tdocs(tdoc_ids: list[str], base_url: str) -> dict[str, str]:
    """Download, extract, and convert the given TDoc IDs.

    Reuses the worker functions from ``download_tdocs`` (no subprocess needed).
    Returns a dict of ``{tdoc_id: status}`` where status is one of
    "downloaded", "extracted", "converted", "skipped", or an error message.
    """
    # Import worker functions from the top-level download script.
    sys.path.insert(0, str(PROJECT_ROOT))
    from download_tdocs import (  # type: ignore[import]
        convert_document_worker,
        download_file_worker,
        extract_file_worker,
    )

    _ensure_artifact_dirs()

    # Temporarily patch the module-level constants used by the workers.
    import download_tdocs as _dt
    _dt.DOWNLOAD_DIR = str(DOWNLOAD_DIR)
    _dt.EXTRACT_DIR = str(EXTRACT_DIR)
    _dt.OUTPUT_DIR = str(ARTIFACTS_DIR / "output")

    results: dict[str, str] = {}

    for tdoc_id in tdoc_ids:
        filename = f"{tdoc_id}.zip"

        # Phase 1 – download
        dl = download_file_worker(filename)
        if dl["status"] == "fail":
            results[tdoc_id] = f"download failed: {dl['message']}"
            continue

        # Phase 2 – extract
        ex = extract_file_worker(filename)
        if ex["status"] == "fail":
            results[tdoc_id] = f"extract failed: {ex['message']}"
            continue

        # Phase 3 – convert all documents inside the extracted folder
        extract_path = EXTRACT_DIR / tdoc_id
        if not extract_path.exists():
            results[tdoc_id] = "extract path missing"
            continue

        converted_any = False
        supported = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls"}
        for root, _, files in (extract_path).walk() if hasattr(extract_path, "walk") else _walk(extract_path):
            for file in files:
                if Path(file).suffix.lower() in supported:
                    file_path = root / file if isinstance(root, Path) else Path(root) / file
                    cv = convert_document_worker((str(file_path), tdoc_id, file))
                    if cv["status"] in ("success", "skip"):
                        converted_any = True

        results[tdoc_id] = "converted" if converted_any else "no supported files found"

    return results


def _walk(path: Path):
    """Fallback os.walk wrapper that yields (Path, dirs, files) tuples."""
    import os
    for root, dirs, files in os.walk(path):
        yield Path(root), dirs, files
